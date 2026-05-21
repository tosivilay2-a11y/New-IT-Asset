"""Build asset import Excel workbook with ref sheets and FK dropdowns."""
import io
from typing import Any, List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .asset_import_schema import ASSET_DB_COLUMNS, FK_DROPDOWN_COLUMNS, INSTRUCTIONS_ROWS


def build_asset_import_workbook(
    sample_row: List[Any],
    reference_sheets: List[Tuple[str, List[str], List[List[Any]]]],
) -> bytes:
    """
    reference_sheets: list of (sheet_name, headers, rows)
    Headers must include 'pick_list' column (used for Excel dropdowns).
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws_inst = wb.create_sheet("Instructions", 0)
    for row in INSTRUCTIONS_ROWS:
        ws_inst.append(row)

    ws = wb.create_sheet("Import", 1)
    ws.append(ASSET_DB_COLUMNS)
    if sample_row:
        ws.append(sample_row)

    ref_pick_columns = {}
    for sheet_name, headers, rows in reference_sheets:
        ref = wb.create_sheet(sheet_name[:31])
        ref.append(headers)
        for r in rows:
            ref.append(r)
        if "pick_list" in headers:
            pick_idx = headers.index("pick_list") + 1
            ref_pick_columns[sheet_name] = get_column_letter(pick_idx)

    for _field, (col_idx, ref_sheet) in FK_DROPDOWN_COLUMNS.items():
        pick_col = ref_pick_columns.get(ref_sheet)
        if not pick_col:
            continue
        col_letter = get_column_letter(col_idx)
        formula = f"='{ref_sheet}'!${pick_col}$2:${pick_col}$500"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Select from list or enter numeric ID from ref sheet"
        dv.errorTitle = "Invalid value"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}2000")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
