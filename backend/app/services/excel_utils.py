"""Shared helpers for Excel/CSV import and export."""
import csv
import io
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def normalize_key(key: Any) -> str:
    if key is None:
        return ""
    return str(key).strip().lower().replace(" ", "").replace("_", "").rstrip("*")


def normalize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    return {normalize_key(k): v for k, v in row.items() if k is not None}


def get_cell(row: Dict[str, Any], *keys: str) -> Any:
    for key in keys:
        val = row.get(normalize_key(key))
        if val is not None and str(val).strip() != "":
            return val
    return None


def parse_upload_file(content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Parse CSV or Excel upload into list of row dicts."""
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    if lower.endswith((".xlsx", ".xls")):
        if not OPENPYXL_AVAILABLE:
            raise ValueError(
                "Excel support not installed. Use CSV or install openpyxl."
            )
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb["Import"] if "Import" in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter, [])]
        rows = []
        for row in rows_iter:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_dict = {}
            for i, header in enumerate(headers):
                if header:
                    row_dict[header] = row[i]
            rows.append(row_dict)
        wb.close()
        return rows

    raise ValueError("Unsupported file format. Use .csv, .xlsx, or .xls")


def build_xlsx(headers: List[str], rows: List[List[Any]], sheet_title: str = "Sheet1") -> bytes:
    """Build an Excel file in memory and return bytes."""
    if not OPENPYXL_AVAILABLE:
        raise ValueError("openpyxl is required for Excel export")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_xlsx_workbook(sheets: List[tuple]) -> bytes:
    """
    Build multi-sheet Excel. Each item: (sheet_title, headers, rows).
  """
    if not OPENPYXL_AVAILABLE:
        raise ValueError("openpyxl is required for Excel export")

    wb = Workbook()
    wb.remove(wb.active)
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for row in rows:
            ws.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_csv(headers: List[str], rows: List[List[Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")
